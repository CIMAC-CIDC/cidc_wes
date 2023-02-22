#!/usr/bin/env python
import argparse
import os
import subprocess
from openpyxl import load_workbook
from subprocess import PIPE, STDOUT


# parse arguments
parser = argparse.ArgumentParser(description="Ingest several samples from google bucket at once")
parser.add_argument("-d", required=True, help="directory containing ingestion templates (required)")
args = parser.parse_args()
directory = args.d
print("directory:", directory)


successes = []
failures = []
for filename in os.listdir(directory):
    file_path = os.path.join(directory, filename)
    print(f"ingesting {file_path} ...")
    workbook = load_workbook(file_path)
    if "WES Analysis" in workbook.sheetnames: # tumor-normal
        assay = "wes_analysis"
    else: # tumor-only
        assay = "wes_tumor_only_analysis"

    result = subprocess.run(f"yes | cidc analyses upload --analysis {assay} --xlsx {file_path}",
                            shell=True, stdout=PIPE, stderr=STDOUT, check=False)
    print(result)
    print(type(result))
    result = result.stdout.decode("utf-8").strip()
    print(result)

    # on success, create new directory needed and move template
    if "Upload succeeded. Visit the CIDC Portal file browser to view your upload." in result:
        successes.append(filename)
        completed_directory = f"{directory}_completed"
        if not os.path.exists(completed_directory):
            os.makedirs(completed_directory)
        os.rename(file_path, os.path.join(completed_directory, filename))
    else:
        failures.append(filename)

    print('-' * 85) # formatting!
    

# final breakdown
print("The following templates were ingested:")
for file in successes:
    print(file)

print("\nThe following templates were not ingested")
for file in failures:
    print(file)
